import requests
import openpyxl
from datetime import datetime

# GitHub API endpoint
base_url = "https://api.github.com"

# Your GitHub personal access token
token = "YOUR_GITHUB_ACCESS_TOKEN"

# Load the Excel workbook
excel_file = "github_users.xlsx"
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Specify the date for which you want to get commits (in YYYY-MM-DD format)
date_to_check = "2023-10-01"

# Create a new Excel workbook to store the results
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# Add headers to the output sheet
output_sheet.append(["GitHub Username", f"Commits on {date_to_check}"])

# Iterate through rows in the Excel sheet
for row in sheet.iter_rows(min_row=2, values_only=True, max_col=1):
    username = row[0]

    # Construct the API URL for user's commits on the specified date
    commits_url = f"{base_url}/users/{username}/events"
    params = {
        "per_page": 100,  # Adjust as needed
        "access_token": token,
    }

    response = requests.get(commits_url, params=params)

    if response.status_code == 200:
        events = response.json()
        commit_count = sum(1 for event in events if event.get("type") == "PushEvent" and event.get("created_at").startswith(date_to_check))
        print(f"User: {username}, Commits on {date_to_check}: {commit_count}")
        output_sheet.append([username, commit_count])
    else:
        print(f"Failed to fetch data for user: {username}, Status Code: {response.status_code}")

# Save the output Excel workbook
output_workbook.save("github_commit_counts.xlsx")
